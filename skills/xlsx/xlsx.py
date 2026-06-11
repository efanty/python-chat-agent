"""XLSX skill — read, create, edit Excel spreadsheets."""

import os
import json
from pathlib import Path


def run(expression: str = "", action: str = "", **kwargs) -> str:
    """Top-level entry point called by SkillExecutor.

    Actions:
      read    — read an .xlsx file and return data as JSON
      create  — create a new spreadsheet
      edit    — edit an existing spreadsheet
      csv     — convert CSV/TSV to xlsx
    """
    action = kwargs.get("action", action) or "read"
    # 所有参数直接通过 kwargs 传入，不需要嵌套的 data/params 结构
    data = kwargs

    if action == "read":
        return _read_xlsx(data)
    elif action == "create":
        return _create_xlsx(data)
    elif action == "edit":
        return _edit_xlsx(data)
    elif action == "csv":
        return _csv_to_xlsx(data)
    else:
        return json.dumps({"success": False, "error": f"未知操作: {action}"},
                          ensure_ascii=False)


def _resolve_path(file_path: str, sandbox_dir: str = None) -> Path:
    """解析文件路径。
    
    如果提供了 sandbox_dir，优先使用沙箱目录。
    否则相对于项目根目录解析。
    """
    p = Path(file_path)
    if p.is_absolute():
        return p
    if sandbox_dir:
        return Path(sandbox_dir) / file_path
    return Path(__file__).resolve().parent.parent.parent / file_path


def _read_xlsx(data: dict) -> str:
    file_path = data.get("file_path", "")
    sheet = data.get("sheet", data.get("sheet_name", None))
    sandbox_dir = data.get("sandbox_dir")
    if not file_path:
        return json.dumps({"success": False, "error": "请提供文件路径"},
                          ensure_ascii=False)
    try:
        import pandas as pd
        path = _resolve_path(file_path, sandbox_dir)
        if not path.exists():
            return json.dumps({"success": False, "error": f"文件不存在: {file_path}"},
                              ensure_ascii=False)
        if sheet:
            df = pd.read_excel(str(path), sheet_name=sheet, dtype=str)
        else:
            df = pd.read_excel(str(path), sheet_name=None, dtype=str)
            result = {}
            for name, sheet_df in df.items():
                result[name] = {
                    "columns": list(sheet_df.columns),
                    "rows": sheet_df.fillna("").to_dict(orient="records"),
                    "shape": list(sheet_df.shape),
                }
            return json.dumps({"success": True, "file": path.name, "sheets": result},
                              ensure_ascii=False)
        return json.dumps({
            "success": True,
            "file": path.name,
            "sheet": sheet if isinstance(sheet, str) else str(sheet),
            "columns": list(df.columns),
            "rows": df.fillna("").to_dict(orient="records"),
            "shape": list(df.shape),
        }, ensure_ascii=False)
    except ImportError:
        return json.dumps({"success": False, "error": "需要 pandas: pip install pandas openpyxl"},
                          ensure_ascii=False)
    except Exception as e:
        return json.dumps({"success": False, "error": f"读取失败: {e}"},
                          ensure_ascii=False)


def _create_xlsx(data: dict) -> str:
    # 兼容 SKILL.md 中定义的参数名和实际代码使用的参数名
    rows = data.get("rows", data.get("content", []))
    columns = data.get("columns", [])
    output = data.get("output", data.get("filename", data.get("output_path", "output.xlsx")))
    sheet_name = data.get("sheet_name", "Sheet1")
    sandbox_dir = data.get("sandbox_dir")
    
    # 如果 rows 是字符串（CSV格式），尝试解析
    if isinstance(rows, str):
        lines = [line.strip() for line in rows.strip().split("\n") if line.strip()]
        if lines:
            # 第一行可能是表头
            first_line = lines[0]
            if "," in first_line or "\t" in first_line:
                import csv
                import io
                reader = csv.reader(io.StringIO(rows))
                all_rows = list(reader)
                if all_rows:
                    if not columns:
                        columns = all_rows[0]
                        rows = all_rows[1:]
                    else:
                        rows = all_rows
    
    if not rows:
        return json.dumps({"success": False, "error": "请提供数据 (rows)"},
                          ensure_ascii=False)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Header
        if columns:
            for c, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=c, value=col_name)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

        # Data
        for r, row_data in enumerate(rows, 2 if columns else 1):
            if isinstance(row_data, dict) and columns:
                for c, col_name in enumerate(columns, 1):
                    ws.cell(row=r, column=c, value=row_data.get(col_name, ""))
            elif isinstance(row_data, (list, tuple)):
                for c, val in enumerate(row_data, 1):
                    ws.cell(row=r, column=c, value=val)

        # Auto-width
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        out_path = _resolve_path(output, sandbox_dir)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(out_path))
        return json.dumps({"success": True, "output": str(out_path), "rows": len(rows),
                           "columns": len(columns) if columns else 0},
                          ensure_ascii=False)
    except ImportError:
        return json.dumps({"success": False, "error": "需要 openpyxl: pip install openpyxl"},
                          ensure_ascii=False)
    except Exception as e:
        return json.dumps({"success": False, "error": f"创建失败: {e}"},
                          ensure_ascii=False)


def _edit_xlsx(data: dict) -> str:
    file_path = data.get("file_path", "")
    output = data.get("output", "edited.xlsx")
    updates = data.get("updates", {})
    sandbox_dir = data.get("sandbox_dir")
    if not file_path:
        return json.dumps({"success": False, "error": "请提供文件路径"},
                          ensure_ascii=False)
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font

        path = _resolve_path(file_path, sandbox_dir)
        if not path.exists():
            return json.dumps({"success": False, "error": f"文件不存在: {file_path}"},
                              ensure_ascii=False)
        wb = load_workbook(str(path))
        for sheet_name, cells in updates.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for cell_ref, value in cells.items():
                ws[cell_ref] = value

        out_path = _resolve_path(output, sandbox_dir)
        wb.save(str(out_path))
        return json.dumps({"success": True, "output": str(out_path)},
                          ensure_ascii=False)
    except ImportError:
        return json.dumps({"success": False, "error": "需要 openpyxl: pip install openpyxl"},
                          ensure_ascii=False)
    except Exception as e:
        return json.dumps({"success": False, "error": f"编辑失败: {e}"},
                          ensure_ascii=False)


def _csv_to_xlsx(data: dict) -> str:
    file_path = data.get("file_path", "")
    output = data.get("output", "")
    delimiter = data.get("delimiter", ",")
    sandbox_dir = data.get("sandbox_dir")
    if not file_path:
        return json.dumps({"success": False, "error": "请提供 CSV 文件路径"},
                          ensure_ascii=False)
    try:
        import pandas as pd
        path = _resolve_path(file_path, sandbox_dir)
        if not path.exists():
            return json.dumps({"success": False, "error": f"文件不存在: {file_path}"},
                              ensure_ascii=False)
        df = pd.read_csv(str(path), delimiter=delimiter)
        if not output:
            output = path.with_suffix(".xlsx").name
        out_path = _resolve_path(output, sandbox_dir)
        df.to_excel(str(out_path), index=False)
        return json.dumps({"success": True, "output": str(out_path),
                           "rows": len(df), "columns": len(df.columns)},
                          ensure_ascii=False)
    except ImportError:
        return json.dumps({"success": False, "error": "需要 pandas: pip install pandas openpyxl"},
                          ensure_ascii=False)
    except Exception as e:
        return json.dumps({"success": False, "error": f"转换失败: {e}"},
                          ensure_ascii=False)
