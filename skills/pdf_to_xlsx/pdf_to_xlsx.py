"""pdf_to_xlsx Skill — 使用智谱视觉大模型将 PDF 表格转换为 Excel 文件。

工作流程:
  1. PDF → 图片 (PyMuPDF/fitz)
  2. 图片 → 结构化表格数据 (智谱 GLM-4V 视觉模型)
  3. 结构化数据 → .xlsx 文件 (openpyxl)
"""

import os
import re
import json
import base64
import uuid
import requests
from pathlib import Path
from io import BytesIO

# ── 路径 ──────────────────────────────────────────────────────────────────

_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
_ENV_PATH = _PROJECT_ROOT / ".env"

# 加载 .env
try:
    from dotenv import load_dotenv
    load_dotenv(_ENV_PATH)
except ImportError:
    pass

# ── API 配置 ──────────────────────────────────────────────────────────────

API_URL = "https://open.bigmodel.cn/api/paas/v4/chat/completions"
DEFAULT_MODEL = os.getenv("ZHIPU_VISION_MODEL", "glm-4v-flash")
FALLBACK_MODELS = ["glm-4v-plus", "glm-5v-turbo"]

# ── 默认提示词 ────────────────────────────────────────────────────────────

DEFAULT_TABLE_PROMPT = (
    "请识别此图片中的所有表格数据。\n"
    "要求：\n"
    "1. 以 JSON 格式输出，格式为 {\"columns\": [\"列名1\", \"列名2\", ...], \"rows\": [{\"列名1\": \"值\", \"列名2\": \"值\"}, ...]}\n"
    "2. 第一行通常是表头，请作为 columns\n"
    "3. 每一行数据作为一个对象，键为列名，值为单元格内容\n"
    "4. 如果有多张表格，请分别输出\n"
    "5. 保留原始数据的数字、日期等格式\n"
    "6. 如果无法识别为表格，请输出 {\"type\": \"text\", \"content\": \"...\"}\n"
    "7. 只输出 JSON，不要其他说明文字"
)


# ── 核心函数 ──────────────────────────────────────────────────────────────

def _resolve_path(file_path: str, sandbox_dir: str = None) -> Path:
    """解析文件路径（相对于项目根目录或绝对路径）。"""
    p = Path(file_path)
    if p.is_absolute():
        return p
    if sandbox_dir:
        return Path(sandbox_dir) / file_path
    return _PROJECT_ROOT / file_path


def _parse_page_range(pages_spec: str, total_pages: int) -> list[int]:
    """解析页码范围，如 '1-3,5,7-9' → [1,2,3,5,7,8,9]。"""
    if not pages_spec or pages_spec.lower() == "all":
        return list(range(1, total_pages + 1))
    pages = []
    for part in pages_spec.split(","):
        part = part.strip()
        if "-" in part:
            start, end = part.split("-", 1)
            pages.extend(range(int(start.strip()), int(end.strip()) + 1))
        else:
            pages.append(int(part))
    return sorted(set(pages))


def _pdf_to_images(pdf_path: Path, dpi: int = 200) -> list[BytesIO]:
    """将 PDF 每页转为 PNG 图片，返回 BytesIO 列表。

    使用 PyMuPDF (fitz) 实现，无需安装 poppler。
    """
    try:
        import fitz  # PyMuPDF
    except ImportError:
        raise ImportError("需要 PyMuPDF: pip install PyMuPDF")

    doc = fitz.open(str(pdf_path))
    zoom = dpi / 72  # PDF 默认 72 DPI
    matrix = fitz.Matrix(zoom, zoom)
    bufs = []
    for page_num in range(len(doc)):
        page = doc[page_num]
        pix = page.get_pixmap(matrix=matrix)
        img_bytes = pix.tobytes("png")
        buf = BytesIO(img_bytes)
        buf.seek(0)
        bufs.append(buf)
    doc.close()
    return bufs


def _encode_image(image_buf: BytesIO) -> str:
    """将图片 BytesIO 编码为 base64 data URL。"""
    b64 = base64.b64encode(image_buf.getvalue()).decode("utf-8")
    return f"data:image/png;base64,{b64}"


def _call_vision_api(data_url: str, prompt: str, model: str, api_key: str) -> tuple[bool, str, str]:
    """调用智谱视觉 API，带模型降级回退。"""
    models_to_try = [model] + [m for m in FALLBACK_MODELS if m != model]
    last_error = ""

    for attempt_model in models_to_try:
        payload = {
            "model": attempt_model,
            "messages": [{
                "role": "user",
                "content": [
                    {"type": "image_url", "image_url": {"url": data_url}},
                    {"type": "text", "text": prompt},
                ],
            }],
        }
        headers = {"Authorization": f"Bearer {api_key}"}
        try:
            resp = requests.post(API_URL, headers=headers, json=payload, timeout=120)
            if resp.status_code == 200:
                data = resp.json()
                choices = data.get("choices", [])
                if choices:
                    text = choices[0].get("message", {}).get("content", "")
                    return True, text, attempt_model
                last_error = "API 返回了空的 choices"
            else:
                last_error = f"HTTP {resp.status_code}: {resp.text[:200]}"
        except requests.exceptions.Timeout:
            last_error = "请求超时"
        except requests.exceptions.RequestException as e:
            last_error = f"请求失败: {e}"
        except Exception as e:
            last_error = f"解析错误: {e}"

    return False, last_error, model


def _parse_vision_response(text: str) -> dict:
    """从视觉模型返回的文本中提取 JSON 数据。"""
    # 尝试直接解析
    text = text.strip()
    # 移除可能的 markdown 代码块标记
    text = re.sub(r'^```(?:json)?\s*', '', text)
    text = re.sub(r'\s*```$', '', text)
    text = text.strip()

    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # 尝试在文本中查找 JSON 对象
    match = re.search(r'\{.*\}', text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group())
        except json.JSONDecodeError:
            pass

    # 尝试提取表格格式（Markdown 表格 → JSON）
    lines = text.strip().split("\n")
    table_lines = [l for l in lines if "|" in l]
    if table_lines:
        # 解析 Markdown 表格
        headers = [h.strip() for h in table_lines[0].split("|") if h.strip()]
        rows = []
        for line in table_lines[2:]:  # 跳过表头分隔行
            cells = [c.strip() for c in line.split("|") if c.strip()]
            if cells and len(cells) == len(headers):
                rows.append(dict(zip(headers, cells)))
        if rows:
            return {"columns": headers, "rows": rows}

    # 无法解析，返回原始文本
    return {"type": "text", "content": text}


def _create_xlsx_from_data(tables: list[dict], output_path: Path, sheet_prefix: str = "Sheet") -> str:
    """将识别到的表格数据写入 Excel 文件。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError("需要 openpyxl: pip install openpyxl")

    wb = Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for idx, table in enumerate(tables):
        columns = table.get("columns", [])
        rows = table.get("rows", [])

        if not columns and not rows:
            continue

        sheet_name = f"{sheet_prefix}_{idx + 1}" if len(tables) > 1 else sheet_prefix
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name max 31 chars

        # 写表头
        for c, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=c, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写数据
        for r, row_data in enumerate(rows, 2):
            if isinstance(row_data, dict):
                for c, col_name in enumerate(columns, 1):
                    cell = ws.cell(row=r, column=c, value=row_data.get(col_name, ""))
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")
            elif isinstance(row_data, (list, tuple)):
                for c, val in enumerate(row_data, 1):
                    cell = ws.cell(row=r, column=c, value=val)
                    cell.border = thin_border

        # 自动列宽
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    cell_len = len(str(cell.value))
                    # 中文字符算 2 个宽度
                    cjk_count = sum(1 for ch in str(cell.value) if '\u4e00' <= ch <= '\u9fff')
                    cell_len += cjk_count
                    max_len = max(max_len, cell_len)
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 60)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return str(output_path)


# ── 入口函数 ──────────────────────────────────────────────────────────────

def run(expression: str = "", action: str = "", **kwargs) -> str:
    """Top-level entry point called by SkillExecutor.

    将 PDF 文件中的表格转换为 Excel 文件。

    Args:
        expression: JSON 字符串或文件路径
        action: "convert"（默认）
        **kwargs:
            file_path: PDF 文件路径（必填）
            output: 输出 Excel 文件名（可选）
            pages: 页码范围，如 "1-3,5" 或 "all"（可选）
            prompt: 自定义识别提示词（可选）
            model: 视觉模型名称（可选）

    Returns:
        JSON 字符串
    """
    action = action or "convert"

    # 解析 JSON expression
    if expression and expression.strip().startswith("{"):
        try:
            args = json.loads(expression)
            file_path = args.get("file_path", kwargs.get("file_path", ""))
            output = args.get("output", kwargs.get("output", ""))
            pages = args.get("pages", kwargs.get("pages", "all"))
            prompt = args.get("prompt", kwargs.get("prompt", ""))
            model = args.get("model", kwargs.get("model", DEFAULT_MODEL))
        except json.JSONDecodeError:
            file_path = expression
            output = kwargs.get("output", "")
            pages = kwargs.get("pages", "all")
            prompt = kwargs.get("prompt", "")
            model = kwargs.get("model", DEFAULT_MODEL)
    else:
        file_path = kwargs.get("file_path", "") or expression
        output = kwargs.get("output", "")
        pages = kwargs.get("pages", "all")
        prompt = kwargs.get("prompt", "")
        model = kwargs.get("model", DEFAULT_MODEL)

    if not file_path:
        return json.dumps({
            "success": False,
            "error": "请提供 PDF 文件路径。用法: run(file_path='/path/to/file.pdf')",
        }, ensure_ascii=False)

    # ── 解析路径 ──────────────────────────────────────────────────────
    pdf_path = _resolve_path(file_path)
    if not pdf_path.exists():
        return json.dumps({
            "success": False,
            "error": f"PDF 文件不存在: {file_path}",
        }, ensure_ascii=False)

    if pdf_path.suffix.lower() != ".pdf":
        return json.dumps({
            "success": False,
            "error": f"不是 PDF 文件: {pdf_path.name}",
        }, ensure_ascii=False)

    # ── 检查 API Key ──────────────────────────────────────────────────
    api_key = os.getenv("ZHIPU_API_KEY", "")
    if not api_key:
        return json.dumps({
            "success": False,
            "error": "ZHIPU_API_KEY 未设置。请从 https://open.bigmodel.cn 获取 API Key。",
        }, ensure_ascii=False)

    # ── 确定输出路径 ──────────────────────────────────────────────────
    sandbox_dir = kwargs.get("sandbox_dir")
    if not output:
        output = pdf_path.with_suffix(".xlsx").name
    output_path = _resolve_path(output, sandbox_dir)

    # ── 步骤 1: PDF 转图片 ────────────────────────────────────────────
    try:
        images = _pdf_to_images(pdf_path)
    except ImportError as e:
        return json.dumps({"success": False, "error": str(e)}, ensure_ascii=False)
    except Exception as e:
        return json.dumps({
            "success": False,
            "error": f"PDF 转图片失败: {e}",

        }, ensure_ascii=False)

    total_pages = len(images)
    target_pages = _parse_page_range(pages, total_pages)

    # ── 步骤 2: 逐页识别 ──────────────────────────────────────────────
    use_prompt = prompt or DEFAULT_TABLE_PROMPT
    all_tables = []
    errors = []

    for page_num in target_pages:
        if page_num < 1 or page_num > total_pages:
            errors.append(f"第 {page_num} 页不存在（共 {total_pages} 页）")
            continue

        img_buf = images[page_num - 1]
        data_url = _encode_image(img_buf)

        success, result_text, model_used = _call_vision_api(data_url, use_prompt, model, api_key)
        if not success:
            errors.append(f"第 {page_num} 页识别失败: {result_text}")
            continue

        parsed = _parse_vision_response(result_text)

        # 如果是表格数据
        if "columns" in parsed and "rows" in parsed:
            parsed["page"] = page_num
            parsed["model_used"] = model_used
            all_tables.append(parsed)
        elif parsed.get("type") == "text":
            # 纯文本内容，尝试作为单列表格
            content = parsed.get("content", "")
            if content.strip():
                all_tables.append({
                    "page": page_num,
                    "columns": ["内容"],
                    "rows": [{"内容": line} for line in content.strip().split("\n") if line.strip()],
                    "model_used": model_used,
                })
        else:
            errors.append(f"第 {page_num} 页未能识别出表格结构")

    if not all_tables:
        return json.dumps({
            "success": False,
            "error": "未能从 PDF 中识别出任何表格数据",
            "details": errors,
            "total_pages": total_pages,
        }, ensure_ascii=False)

    # ── 步骤 3: 重建 Excel ────────────────────────────────────────────
    try:
        xlsx_path = _create_xlsx_from_data(all_tables, output_path)
    except Exception as e:
        return json.dumps({
            "success": False,
            "error": f"创建 Excel 文件失败: {e}",
        }, ensure_ascii=False)

    result = {
        "success": True,
        "output": xlsx_path,
        "total_pages": total_pages,
        "processed_pages": len(target_pages),
        "tables_found": len(all_tables),
        "model_used": model,
    }
    if errors:
        result["warnings"] = errors

    return json.dumps(result, ensure_ascii=False)


# ── Action aliases ────────────────────────────────────────────────────────

convert = run
